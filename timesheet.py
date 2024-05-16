from openpyxl import load_workbook
from datetime import datetime, timedelta
import calendar

def is_weekend(date):
    # Monday is 0 and Sunday is 6
    return date.weekday() >= 5

def generate_dates():
    today = datetime.today()
    first_day_of_month = today.replace(day=1)
    next_month = first_day_of_month.replace(month=first_day_of_month.month + 1)
    last_day_of_month = next_month - timedelta(days=1)
    current_date = first_day_of_month
    while current_date <= last_day_of_month:
        yield current_date
        current_date += timedelta(days=1)

def modify_xlsx(file_path, sheet_name,public_holiday=[],annual_leave=[],sick_leave=[]):
    wb = load_workbook(filename=file_path)
    ws = wb[sheet_name]
    
    #Generate dates
    dates = generate_dates()
    
    #Delete last month timesheet
    try:
       start_cell = 'A11'
       end_cell = 'H41'
       cell_range = ws[start_cell:end_cell]
       for row in cell_range:
          for cell in row:
             cell.value = '' 
    except IndexError:
        print("Error: Row or column index out of range.")
        return
    
    #Add timesheet for current month
    try:
        row_index=11
        workdays=0
        for date in dates:
           ws.cell(row=row_index, column=1, value=str(row_index-10))
           ws.cell(row=row_index, column=2, value=date)
           if date.day in public_holiday:
              ws.cell(row=row_index, column=3, value="-")
              ws.cell(row=row_index, column=4, value="1")
           elif date.day in annual_leave:
              ws.cell(row=row_index, column=3, value="-")
              ws.cell(row=row_index, column=7, value="1")
           elif date.day in sick_leave:
              ws.cell(row=row_index, column=3, value="-")
              ws.cell(row=row_index, column=5, value="1")
           elif is_weekend(date):
               ws.cell(row=row_index, column=3, value="-")
               ws.cell(row=row_index, column=8, value="Weekend")
           else:
               ws.cell(row=row_index, column=3, value="1")
               workdays=workdays+1
           row_index=row_index+1
        month_name = calendar.month_name[date.month]
        year=date.year
        ws.cell(row=43, column=3, value=workdays)
        ws.cell(row=2, column=7, value='-'.join([month_name,str(year)]))
        ws.cell(row=46, column=2, value=date)
        ws.cell(row=50, column=2, value=date)
    except IndexError:
        print("Error: Row or column index out of range.")
        return
    wb.save(file_path)
    print("Timesheet generated successfully.")

file_path = 'timesheet.xlsx'
sheet_name = 'Timesheet'
public_holiday=[1,22]
annual_leave=[]
sick_leave=[]

modify_xlsx(file_path, sheet_name,public_holiday,annual_leave,sick_leave)
